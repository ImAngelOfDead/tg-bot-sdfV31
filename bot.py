import os
import logging
import asyncio
import psycopg2
import re
import csv 
import io

from datetime import datetime, timedelta, date
from io import BytesIO
from dotenv import load_dotenv
from openpyxl import Workbook
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
from aiogram.filters import Command 
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from openpyxl.styles import Alignment, Font

load_dotenv()   

# Конфигурация подключения к базе данных
DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

# Токен бота
TOKEN = os.getenv("TOKEN")

# Логирование
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(level=LOG_LEVEL)

# Параметры расчёта дат
MIN_DATE_OFFSET = int(os.getenv("MIN_DATE_OFFSET", 2))
MAX_DATE_OFFSET = int(os.getenv("MAX_DATE_OFFSET", 30))

# Параметры клавиатуры
PAGE_SIZE = int(os.getenv("PAGE_SIZE", 5))

# Тексты сообщений и кнопок
TEXT_WELCOME = os.getenv("TEXT_WELCOME", "Привет! Выбери действие:")
TEXT_MENU = os.getenv("TEXT_MENU", "Главное меню")
TEXT_ADMIN_REQUIRED = os.getenv("TEXT_ADMIN_REQUIRED", "Требуются права администратора")
TEXT_INVALID_DATE_FORMAT = os.getenv("TEXT_INVALID_DATE_FORMAT", "Неверный формат даты. Используйте ДД.ММ.ГГГГ")
TEXT_INVALID_PERIOD = os.getenv("TEXT_INVALID_PERIOD", "Некорректный временной период")
BUTTON_WORK_TIME = os.getenv("BUTTON_WORK_TIME", "Время работы")
BUTTON_DAY_OFF = os.getenv("BUTTON_DAY_OFF", "Поставить выходной")
BUTTON_START_SHIFT = os.getenv("BUTTON_START_SHIFT", "Начать смену")
BUTTON_START_BREAK = os.getenv("BUTTON_START_BREAK", "Начать перерыв")
BUTTON_END_SHIFT = os.getenv("BUTTON_END_SHIFT", "Закончить смену")
BUTTON_END_BREAK = os.getenv("BUTTON_END_BREAK", "Закончить перерыв")
BUTTON_GET_REPORT = os.getenv("BUTTON_GET_REPORT", "Сформировать отчет")

# Callback данные для inline кнопок
CALLBACK_CONFIRM_END_SHIFT = os.getenv("CALLBACK_CONFIRM_END_SHIFT", "confirm_end_shift")
CALLBACK_CANCEL_END_SHIFT = os.getenv("CALLBACK_CANCEL_END_SHIFT", "cancel_end_shift")
CALLBACK_CONFIRM_END_BREAK = os.getenv("CALLBACK_CONFIRM_END_BREAK", "confirm_end_break")
CALLBACK_CANCEL_END_BREAK = os.getenv("CALLBACK_CANCEL_END_BREAK", "cancel_end_break")

# Операционные команды
OPERATION_START_SHIFT = os.getenv("OPERATION_START_SHIFT", "start_shift")
OPERATION_END_SHIFT = os.getenv("OPERATION_END_SHIFT", "end_shift")
OPERATION_START_BREAK = os.getenv("OPERATION_START_BREAK", "start_break")
OPERATION_END_BREAK = os.getenv("OPERATION_END_BREAK", "end_break")
OPERATION_PHOTO_RECEIVED = os.getenv("OPERATION_PHOTO_RECEIVED", "photo_received")

# Регулярное выражение для проверки даты
DATE_REGEX = os.getenv("DATE_REGEX", r"\d{2}\.\d{2}\.\d{4}")

# Инициализация бота и диспетчера
bot = Bot(token=TOKEN)
dp = Dispatcher()

# Установка соединения с базой данных
conn = psycopg2.connect(host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASSWORD)
cursor = conn.cursor()

# Создание таблиц, если они ещё не созданы
cursor.execute("""
CREATE TABLE IF NOT EXISTS users (
    id SERIAL PRIMARY KEY,
    full_name VARCHAR,
    telegram_id VARCHAR UNIQUE,
    department VARCHAR,
    position VARCHAR,
    is_admin BOOLEAN DEFAULT FALSE,
    reminder VARCHAR
);
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS weekends (
    user_id INTEGER NOT NULL,
    date DATE NOT NULL,
    CONSTRAINT fk_weekends_user
        FOREIGN KEY (user_id)
        REFERENCES users (id)
        ON DELETE CASCADE
);
""")
cursor.execute("""
CREATE TABLE IF NOT EXISTS operations (
    user_id INTEGER NOT NULL,
    operation VARCHAR NOT NULL,
    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    CONSTRAINT fk_operations_user
        FOREIGN KEY (user_id)
        REFERENCES users (id)
        ON DELETE CASCADE
);
""")
conn.commit()

def format_time(dt):
    return dt.strftime("%d.%m.%Y %H:%M:%S") if dt else ""

def get_or_create_user(telegram_id: str) -> int:
    cursor.execute("SELECT id FROM users WHERE telegram_id = %s", (telegram_id,))
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute("INSERT INTO users (telegram_id) VALUES (%s) RETURNING id", (telegram_id,))
    new_id = cursor.fetchone()[0]
    conn.commit()
    return new_id

def insert_operation(user_id: int, operation: str):
    cursor.execute("INSERT INTO operations (user_id, operation) VALUES (%s, %s)", (user_id, operation))
    conn.commit()

def get_last_operation_time(user_id: int, operation: str):
    cursor.execute("""
        SELECT created_at 
        FROM operations
        WHERE user_id = %s AND operation = %s
        ORDER BY created_at DESC
        LIMIT 1
    """, (user_id, operation))
    row = cursor.fetchone()
    return row[0] if row else None

def is_shift_active(user_id: int) -> bool:
    start_time = get_last_operation_time(user_id, OPERATION_START_SHIFT)
    if not start_time:
        return False
    cursor.execute("""
        SELECT COUNT(*) 
        FROM operations
        WHERE user_id = %s AND operation = %s AND created_at > %s
    """, (user_id, OPERATION_END_SHIFT, start_time))
    return cursor.fetchone()[0] == 0

def is_break_active(user_id: int) -> bool:
    start_time = get_last_operation_time(user_id, OPERATION_START_BREAK)
    if not start_time:
        return False
    cursor.execute("""
        SELECT COUNT(*) 
        FROM operations
        WHERE user_id = %s AND operation = %s AND created_at > %s
    """, (user_id, OPERATION_END_BREAK, start_time))
    return cursor.fetchone()[0] == 0

def calculate_break_duration(user_id: int, shift_start: datetime, shift_end: datetime) -> timedelta:
    total_break = timedelta()
    cursor.execute("""
        SELECT operation, created_at 
        FROM operations 
        WHERE user_id = %s 
            AND operation IN (%s, %s) 
            AND created_at BETWEEN %s AND %s
        ORDER BY created_at
    """, (user_id, OPERATION_START_BREAK, OPERATION_END_BREAK, shift_start, shift_end or datetime.now()))
    
    start_break = None
    for op_type, op_time in cursor.fetchall():
        if op_type == OPERATION_START_BREAK:
            start_break = op_time
        elif op_type == OPERATION_END_BREAK and start_break:
            total_break += op_time - start_break
            start_break = None
    return total_break

def get_last_shift_times(user_id: int):
    st_time = get_last_operation_time(user_id, OPERATION_START_SHIFT)
    if not st_time:
        return (None, None)
    cursor.execute("""
        SELECT created_at 
        FROM operations
        WHERE user_id = %s AND operation = %s AND created_at > %s
        ORDER BY created_at ASC
        LIMIT 1
    """, (user_id, OPERATION_END_SHIFT, st_time))
    row = cursor.fetchone()
    return (st_time, row[0]) if row else (st_time, None)

def get_user_reminder(user_id: int):
    cursor.execute("SELECT reminder FROM users WHERE id = %s", (user_id,))
    row = cursor.fetchone()
    return row[0] if row and row[0] else ""

# Словарь для хранения текущей страницы календаря для каждого пользователя
user_dayoff_pages = {}

def build_day_off_inline_keyboard(user_id: int, page_start: datetime.date) -> InlineKeyboardMarkup:
    today = datetime.date.today()
    min_date = today + datetime.timedelta(days=MIN_DATE_OFFSET)
    max_date = today + datetime.timedelta(days=MAX_DATE_OFFSET)
    if page_start < min_date:
        page_start = min_date

    keyboard = InlineKeyboardMarkup(inline_keyboard=[], row_width=1)
    current_date = page_start
    # Добавляем кнопки с доступными датами
    for _ in range(PAGE_SIZE):
        if current_date > max_date:
            break
        cursor.execute("SELECT COUNT(*) FROM weekends WHERE date = %s", (current_date,))
        count = cursor.fetchone()[0]
        if count == 0:
            keyboard.inline_keyboard.append([
                InlineKeyboardButton(
                    text=current_date.strftime("%d.%m.%Y"),
                    callback_data=f"day_off_select:{current_date.strftime('%Y-%m-%d')}"
                )
            ])
        current_date += datetime.timedelta(days=1)

    # Формируем ряд навигационных кнопок
    nav_buttons = []
    if page_start > min_date:
        nav_buttons.append(InlineKeyboardButton(text="←", callback_data="day_off_prev"))
    if current_date <= max_date:
        nav_buttons.append(InlineKeyboardButton(text="→", callback_data="day_off_next"))
    nav_buttons.append(InlineKeyboardButton(text="Назад", callback_data="day_off_back"))
    if nav_buttons:
        keyboard.inline_keyboard.append(nav_buttons)
    return keyboard


@dp.message(lambda msg: msg.text == BUTTON_DAY_OFF)
async def ask_day_off_date(message: types.Message):
    user_id = get_or_create_user(str(message.from_user.id))
    today = datetime.date.today()
    min_date = today + datetime.timedelta(days=MIN_DATE_OFFSET)
    user_dayoff_pages[user_id] = min_date
    kb = build_day_off_inline_keyboard(user_id, min_date)
    await message.answer("Выберите дату для выходного:", reply_markup=kb)

@dp.callback_query(lambda c: c.data.startswith("day_off_select:"))
async def handle_day_off_select(callback_query: types.CallbackQuery):
    user_id = get_or_create_user(str(callback_query.from_user.id))
    date_str = callback_query.data.split(":")[1]
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    # Проверяем доступность даты
    cursor.execute("SELECT COUNT(*) FROM weekends WHERE date = %s", (selected_date,))
    count = cursor.fetchone()[0]
    if count > 0:
        await callback_query.answer("Этот день уже занят.", show_alert=True)
    else:
        cursor.execute("INSERT INTO weekends (user_id, date) VALUES (%s, %s)", (user_id, selected_date))
        conn.commit()
        await callback_query.message.edit_text(f"Выходной на {selected_date.strftime('%d.%m.%Y')} установлен.")
        await bot.send_message(callback_query.from_user.id, TEXT_MENU, reply_markup=menu_keyboard)
    user_dayoff_pages.pop(user_id, None)

@dp.callback_query(lambda c: c.data in ["day_off_prev", "day_off_next", "day_off_back"])
async def day_off_navigation(callback_query: types.CallbackQuery):
    user_id = get_or_create_user(str(callback_query.from_user.id))
    if callback_query.data == "day_off_back":
        user_dayoff_pages.pop(user_id, None)
        await callback_query.message.edit_text(TEXT_MENU, reply_markup=menu_keyboard)
        return

    page_start = user_dayoff_pages.get(user_id)
    if not page_start:
        page_start = datetime.date.today() + datetime.timedelta(days=MIN_DATE_OFFSET)

    today = datetime.date.today()
    if callback_query.data == "day_off_prev":
        new_start = page_start - datetime.timedelta(days=PAGE_SIZE)
        min_date = today + datetime.timedelta(days=MIN_DATE_OFFSET)
        if new_start < min_date:
            new_start = min_date
    elif callback_query.data == "day_off_next":
        new_start = page_start + datetime.timedelta(days=PAGE_SIZE)
        max_date = today + datetime.timedelta(days=MAX_DATE_OFFSET)
        if new_start > max_date:
            new_start = max_date
    user_dayoff_pages[user_id] = new_start
    kb = build_day_off_inline_keyboard(user_id, new_start)
    await callback_query.message.edit_reply_markup(reply_markup=kb)

menu_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=BUTTON_WORK_TIME), KeyboardButton(text=BUTTON_DAY_OFF)],
        [KeyboardButton(text=BUTTON_START_SHIFT), KeyboardButton(text=BUTTON_START_BREAK),
         KeyboardButton(text=BUTTON_END_SHIFT), KeyboardButton(text=BUTTON_END_BREAK)],
        [KeyboardButton(text=BUTTON_GET_REPORT)]
    ],
    resize_keyboard=True
)

confirm_shift_keyboard = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="Подтвердить завершение смены", callback_data=CALLBACK_CONFIRM_END_SHIFT)],
    [InlineKeyboardButton(text="Отмена", callback_data=CALLBACK_CANCEL_END_SHIFT)]
])

confirm_break_keyboard = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="Подтвердить завершение перерыва", callback_data=CALLBACK_CONFIRM_END_BREAK)],
    [InlineKeyboardButton(text="Отмена", callback_data=CALLBACK_CANCEL_END_BREAK)]
])

@dp.message(Command("start"))
async def start_command(message: types.Message):
    await message.answer(TEXT_WELCOME, reply_markup=menu_keyboard)
    
@dp.message(Command("get"))
async def handle_get_report(message: types.Message, state: FSMContext):
    user_id = get_or_create_user(str(message.from_user.id))
    cursor.execute("SELECT is_admin FROM users WHERE id = %s", (user_id,))
    if not cursor.fetchone()[0]:
        await message.answer(TEXT_ADMIN_REQUIRED)
        return

    try:
        args = message.text.split()[1:]
        if len(args) < 2:
            await message.answer("Укажите две даты: ДД.ММ.ГГГГ ДД.ММ.ГГГГ")
            return

        date_from = datetime.strptime(args[0], "%d.%m.%Y").date()
        date_to = datetime.strptime(args[1], "%d.%m.%Y").date()

        if date_from > date_to:
            await message.answer(TEXT_INVALID_PERIOD)
            return

        await state.update_data(date_from=date_from, date_to=date_to)
        await state.set_state(ReportStates.WAITING_FOR_FORMAT)

        format_keyboard = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="CSV")],
                [KeyboardButton(text="Excel")]
            ],
            resize_keyboard=True
        )
        await message.answer("Выберите формат отчёта:", reply_markup=format_keyboard)

    except ValueError:
        await message.answer(TEXT_INVALID_DATE_FORMAT)
        
@dp.message(lambda msg: msg.text == BUTTON_START_SHIFT)
async def start_shift(message: types.Message):
    user_id = get_or_create_user(str(message.from_user.id))
    if is_shift_active(user_id):
        await message.answer("У вас уже есть активная смена. Завершите её.")
        return
    insert_operation(user_id, OPERATION_START_SHIFT)
    start_time = get_last_operation_time(user_id, OPERATION_START_SHIFT)
    await message.answer(f"Смена начата в {format_time(start_time)}. Пришли фото рабочего места, если требуется.")

@dp.message(lambda msg: msg.photo)
async def receive_photo(message: types.Message):
    user_id = get_or_create_user(str(message.from_user.id))
    if not is_shift_active(user_id):
        await message.answer("Нет активной смены для фото.")
        return
    insert_operation(user_id, OPERATION_PHOTO_RECEIVED)
    await message.answer("Фото принято. Хорошей смены!")

@dp.message(lambda msg: msg.text == BUTTON_START_BREAK)
async def start_break(message: types.Message):
    user_id = get_or_create_user(str(message.from_user.id))
    if not is_shift_active(user_id):
        await message.answer("Сначала начните смену.")
        return
    if is_break_active(user_id):
        await message.answer("Перерыв уже идет. Завершите его.")
        return
    insert_operation(user_id, OPERATION_START_BREAK)
    start_time = get_last_operation_time(user_id, OPERATION_START_BREAK)
    await message.answer(f"Перерыв начат в {format_time(start_time)}.")

@dp.message(lambda msg: msg.text == BUTTON_END_BREAK)
async def request_end_break(message: types.Message):
    user_id = get_or_create_user(str(message.from_user.id))
    if not is_shift_active(user_id):
        await message.answer("Нет активной смены.")
        return
    if not is_break_active(user_id):
        await message.answer("Перерыв не начат или уже завершен.")
        return
    await message.answer("Завершить перерыв?", reply_markup=confirm_break_keyboard)

@dp.callback_query(lambda c: c.data == CALLBACK_CONFIRM_END_BREAK)
async def confirm_end_break(callback_query: types.CallbackQuery):
    user_id = get_or_create_user(str(callback_query.from_user.id))
    insert_operation(user_id, OPERATION_END_BREAK)
    end_time = get_last_operation_time(user_id, OPERATION_END_BREAK)
    await callback_query.answer()
    await callback_query.message.edit_text(f"Перерыв завершён в {format_time(end_time)}")
    await bot.send_message(callback_query.from_user.id, TEXT_MENU, reply_markup=menu_keyboard)

@dp.callback_query(lambda c: c.data == CALLBACK_CANCEL_END_BREAK)
async def cancel_end_break(callback_query: types.CallbackQuery):
    await callback_query.answer()
    await callback_query.message.edit_text("Операция отменена.")
    await bot.send_message(callback_query.from_user.id, TEXT_MENU, reply_markup=menu_keyboard)

@dp.message(lambda msg: msg.text == BUTTON_END_SHIFT)
async def request_end_shift(message: types.Message):
    user_id = get_or_create_user(str(message.from_user.id))
    if not is_shift_active(user_id):
        await message.answer("Нет активной смены.")
        return
    reminder_text = get_user_reminder(user_id)
    confirm_text = "Завершить смену?"
    if reminder_text:
        confirm_text += f"\nНапоминание: {reminder_text}"
    await message.answer(confirm_text, reply_markup=confirm_shift_keyboard)

class ReportStates(StatesGroup):
    WAITING_FOR_DATE_FROM = State()
    WAITING_FOR_DATE_TO = State()
    WAITING_FOR_FORMAT = State()

@dp.message(lambda msg: msg.text == BUTTON_GET_REPORT)
async def request_report(message: types.Message, state: FSMContext):
    user_id = get_or_create_user(str(message.from_user.id))
    cursor.execute("SELECT is_admin FROM users WHERE id = %s", (user_id,))
    if not cursor.fetchone()[0]:
        await message.answer(TEXT_ADMIN_REQUIRED)
        return

    await state.set_state(ReportStates.WAITING_FOR_DATE_FROM)
    await message.answer("Введите дату начала отчёта в формате ДД.ММ.ГГГГ:")

# Обработчик ввода даты "от"
@dp.message(ReportStates.WAITING_FOR_DATE_FROM)
async def handle_date_from(message: types.Message, state: FSMContext):
    try:
        date_from = datetime.strptime(message.text, "%d.%m.%Y").date()
        await state.update_data(date_from=date_from)
        await state.set_state(ReportStates.WAITING_FOR_DATE_TO)
        await message.answer("Введите дату окончания отчёта в формате ДД.ММ.ГГГГ:")
    except ValueError:
        await message.answer(TEXT_INVALID_DATE_FORMAT)

# Обработчик ввода даты "до"
@dp.message(ReportStates.WAITING_FOR_DATE_TO)
async def handle_date_to(message: types.Message, state: FSMContext):
    try:
        date_to = datetime.strptime(message.text, "%d.%m.%Y").date()
        data = await state.get_data()
        date_from = data["date_from"]

        if date_from > date_to:
            await message.answer(TEXT_INVALID_PERIOD)
            return

        await state.update_data(date_to=date_to)
        await state.set_state(ReportStates.WAITING_FOR_FORMAT)

        format_keyboard = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="CSV")],
                [KeyboardButton(text="Excel")]
            ],
            resize_keyboard=True
        )
        await message.answer("Выберите формат отчёта:", reply_markup=format_keyboard)
    except ValueError:
        await message.answer(TEXT_INVALID_DATE_FORMAT)

# Обработчик выбора формата
@dp.message(ReportStates.WAITING_FOR_FORMAT)
async def handle_format_choice(message: types.Message, state: FSMContext):
    format_choice = message.text.lower()
    if format_choice not in ["csv", "excel"]:
        await message.answer("Пожалуйста, выберите формат: CSV или Excel.")
        return

    data = await state.get_data()
    date_from = data["date_from"]
    date_to = data["date_to"]

    if format_choice == "csv":
        report_file = await generate_report_csv(date_from, date_to)
        filename = f"report_{date_from}_{date_to}.csv"
    else:
        report_file = await generate_report_excel(date_from, date_to)
        filename = f"report_{date_from}_{date_to}.xlsx"

    await message.answer_document(
        document=types.BufferedInputFile(
            report_file.getvalue(),
            filename=filename
        )
    )

    await state.clear()
    await message.answer("Отчёт отправлен.", reply_markup=types.ReplyKeyboardRemove())
    await message.answer(TEXT_MENU, reply_markup=menu_keyboard)

# Функция генерации CSV-отчёта
async def generate_report_csv(start_date: date, end_date: date) -> BytesIO:
    output = io.StringIO(newline='')
    writer = csv.writer(output, delimiter=';')

    cursor.execute("SELECT id, full_name FROM users")
    for user_id, full_name in cursor.fetchall():
        writer.writerow([full_name])
        writer.writerow(["Дата", "Начало смены", "Конец смены", "Перерывы"])

        cursor.execute("""
            SELECT created_at 
            FROM operations 
            WHERE user_id = %s 
                AND operation = %s 
                AND created_at::date BETWEEN %s AND %s
            ORDER BY created_at
        """, (user_id, OPERATION_START_SHIFT, start_date, end_date))
        
        for shift_start_row in cursor.fetchall():
            shift_start = shift_start_row[0]
            
            cursor.execute("""
                SELECT created_at 
                FROM operations 
                WHERE user_id = %s 
                    AND operation = %s 
                    AND created_at > %s 
                ORDER BY created_at 
                LIMIT 1
            """, (user_id, OPERATION_END_SHIFT, shift_start))
            shift_end_row = cursor.fetchone()
            shift_end = shift_end_row[0] if shift_end_row else None
            
            break_duration = calculate_break_duration(user_id, shift_start, shift_end)
            total_seconds = int(break_duration.total_seconds())
            minutes, seconds = divmod(total_seconds, 60)
            break_duration_str = f"{minutes}:{seconds:02d}"

            writer.writerow([
                shift_start.date().strftime("%d.%m.%Y"),
                shift_start.time().strftime("%H:%M"),
                shift_end.time().strftime("%H:%M") if shift_end else 'Не завершена',
                break_duration_str
            ])

        writer.writerow([])

    output.seek(0)
    return BytesIO(output.getvalue().encode())

# Функция генерации Excel-отчёта
async def generate_report_excel(start_date: date, end_date: date) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    cursor.execute("SELECT id, full_name FROM users")
    for user_id, full_name in cursor.fetchall():
        ws.append([full_name])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
        ws.cell(row=ws.max_row, column=1).font = header_font
        ws.cell(row=ws.max_row, column=1).alignment = center_alignment

        ws.append(["Дата", "Начало смены", "Конец смены", "Перерывы"])
        for col in range(1, 5):
            ws.cell(row=ws.max_row, column=col).font = header_font
            ws.cell(row=ws.max_row, column=col).alignment = center_alignment

        cursor.execute("""
            SELECT created_at 
            FROM operations 
            WHERE user_id = %s 
                AND operation = %s 
                AND created_at::date BETWEEN %s AND %s
            ORDER BY created_at
        """, (user_id, OPERATION_START_SHIFT, start_date, end_date))
        
        for shift_start_row in cursor.fetchall():
            shift_start = shift_start_row[0]
            
            cursor.execute("""
                SELECT created_at 
                FROM operations 
                WHERE user_id = %s 
                    AND operation = %s 
                    AND created_at > %s 
                ORDER BY created_at 
                LIMIT 1
            """, (user_id, OPERATION_END_SHIFT, shift_start))
            shift_end_row = cursor.fetchone()
            shift_end = shift_end_row[0] if shift_end_row else None
            
            break_duration = calculate_break_duration(user_id, shift_start, shift_end)
            
            total_seconds = int(break_duration.total_seconds())
            minutes, seconds = divmod(total_seconds, 60)
            break_duration_str = f"{minutes}:{seconds:02d}"

            ws.append([
                shift_start.date().strftime("%d.%m.%Y"),  
                shift_start.time().strftime("%H:%M"),    
                shift_end.time().strftime("%H:%M") if shift_end else 'Не завершена',  
                break_duration_str                        
            ])

        ws.append([])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@dp.callback_query(lambda c: c.data == CALLBACK_CONFIRM_END_SHIFT)
async def confirm_end_shift(callback_query: types.CallbackQuery):
    user_id = get_or_create_user(str(callback_query.from_user.id))
    insert_operation(user_id, OPERATION_END_SHIFT)
    end_time = get_last_operation_time(user_id, OPERATION_END_SHIFT)
    response_text = f"Смена завершена в {format_time(end_time)}."
    await callback_query.answer()
    await callback_query.message.edit_text(response_text)
    await bot.send_message(callback_query.from_user.id, TEXT_MENU, reply_markup=menu_keyboard)


@dp.callback_query(lambda c: c.data == CALLBACK_CANCEL_END_SHIFT)
async def cancel_end_shift(callback_query: types.CallbackQuery):
    await callback_query.answer()
    await callback_query.message.edit_text("Операция отменена.")
    await bot.send_message(callback_query.from_user.id, TEXT_MENU, reply_markup=menu_keyboard)

@dp.message(lambda msg: msg.text == BUTTON_WORK_TIME)
async def work_time(message: types.Message):
    user_id = get_or_create_user(str(message.from_user.id))
    start_time, end_time = get_last_shift_times(user_id)
    if not start_time:
        await message.answer("Смена не начиналась.")
    else:
        if end_time:
            await message.answer(f"Смена:\nНачало: {format_time(start_time)}\nКонец: {format_time(end_time)}")
        else:
            await message.answer(f"Смена:\nНачало: {format_time(start_time)}\nНе завершена")

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
